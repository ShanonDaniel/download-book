import requests
from yaspin import yaspin
from yaspin.spinners import Spinners
from tkinter import Tk
from tkinter import messagebox
from tkinter import filedialog

from clint.textui import colored, puts

from openpyxl import load_workbook

import os
from sys import platform

from pypdf import PdfReader, PdfWriter
from reportlab.lib.units import inch

from pdf2image import convert_from_path
from PIL import Image

from config import *

def download_books_quantity(total_rows):
    total_rows=1000
    #Tk().withdraw()
    #messagebox.showinfo(title="Input Required",message="Please enter number of books you want to upload")
    puts(colored.cyan("\n(INPUT): [?] How many books do you want to upload?"))
    while(True):
        userInput_books_upload_number = input(">> ")
        try:
            if(userInput_books_upload_number == "" or userInput_books_upload_number == None):
                Tk().withdraw()
                messagebox.showerror(title="Input Error",message="Please provide a number")
            else:
                userInput_books_upload_number = int(userInput_books_upload_number)
                if(userInput_books_upload_number > total_rows):
                    Tk().withdraw()
                    msg = "(Selected) Excel Sheet has "+str(total_rows)+" rows \nYou've selected "+str(userInput_books_upload_number)+" \nInput can't be greater than Total Excel rows"
                    messagebox.showerror(title="Input Error",message=msg)
                else:
                    break
        except:
            Tk().withdraw()
            messagebox.showerror(title="Input Error",message="Seems like you've enter a string NOT a number\nPlease Enter a valid Number")
    userInput_books_upload_number = int(userInput_books_upload_number)
    Tk().withdraw()
    if platform != "darwin":
        messagebox.showinfo(title="Heads-up Notice", message= "[Bot] Hey,\nI'm about to start\nMake sure (selected) Excel Sheet is not open & is not in Read-Only mode")
    # Returns input 
    return userInput_books_upload_number

def get_excelsheetdata():
    with yaspin(text="Select Excel Sheet Data File" ).magenta.bold.blink.bouncingBall as sp:
        while(True):
            Tk().withdraw()
            imported_file_path = filedialog.askopenfilename()
            if not (imported_file_path.endswith(".xlsx")):
                messagebox.showerror(title="Excel File Error", message="You have selected non-supported file type\nPlease select Excel (.xlsx) file")
            else:
                break
            
        sp.ok("✅")
        return(imported_file_path)

def searcher(sheet):

    #for value in sheet.iter_cols(max_col=2,values_only=True):
    #   print (value)
    totalrows = sheet.max_row
    #print(totalrows)
    i = 0
    
    while (i < totalrows):
        cell_add= "A"+str(i + 2)
        cell = sheet[cell_add]
        cell_value = str(cell.value)
        if (cell_value == 'None' or cell_value == "" ):
            return str(cell)            
        else:
            pass
        i += 1
    
    return None

def getrow(cell_add,sheet,workbook_handle,sheet_name):
    
    # replace query
    q1 = "<Cell '"+sheet_name+"'.A"
    
    # Replace junks chars from the cell_add string
    cell_add = cell_add.replace(q1,"")
    cell_add = cell_add.replace(">","")

    # row_index contains EXTRACTED ROW NUMBER
    row_index = int(cell_add)
    # print(f'index is {row_index}')           ### DEBUG
    
    # assigning row_id to GLOBAL pd_var
    row_id_PD = row_index
    
    # Gets the row via row_index            e.g Row # 10
    row = sheet[row_index]

    
    # Replaces None with ""
    for i in row:
        if(i.value == None):
            i.value = ""
            #print(i.value)

    return(row)

def download_file(url, filename):
    try:
        os.mkdir(f'download/{filename}')
        puts(colored.green(f'Success to make a {filename} directory!'))
    except:
        puts(colored.yellow(f'Already exist a {filename} directory!'))
    
    # NOTE the stream=True parameter below
    with yaspin(text="Downloading a pdf file...", color="cyan") as sp:
        try:
            with requests.get(url, stream=True) as r:
                r.raise_for_status()
                with open(f'download/{filename}/tmp_{filename}.pdf', 'wb') as f:
                    for chunk in r.iter_content(chunk_size=1024 * 8):
                        if chunk:
                            f.write(chunk)
                            f.flush()
                            os.fsync(f.fileno())
                sp.write(f"✅ {filename} download complete!")
                return True
        except:
            sp.write(f"❌ {filename} download Failed!")
            return False

if __name__ == "__main__":
    # Make download directory
    try:
        os.mkdir('download')
        puts(colored.green('Success to make a download directory!'))
    except:
        print(colored.yellow('Already exist a download directory!'))
    Excelfilepath= get_excelsheetdata()
    
    try:
        with yaspin(text="Gathering Data from your Excel File", color = "green") as sp:
            workbook = load_workbook(filename=Excelfilepath)

            #Sheet handle : makes the sheet active so we can interact with sheet's data
            sheet = workbook.active
    
            #gets name of sheets in the workbook    
            sheets = workbook.sheetnames
    
            # contains name of the first sheet
            sheet_1 = sheets[0]
            
            #calculates total number of rows in the Provided Excel sheet column
            total_rows = int(sheet.max_row)-1
            sp.write("✅ Finished reading Excel file!")
    except:
        puts(colored.cyan("\n[ERROR] : Provide Excel is CORRUPTED - Script can't open it"))
    
    userInput_books_download_number = download_books_quantity(0)

    books_count = 0

    while(books_count < userInput_books_download_number):
        row_id=searcher(sheet)
        
        ebookDetails = getrow(str(row_id),sheet,workbook,sheet_1)
        
        if download_file(ebookDetails[2].value, ebookDetails[1].value):
            ebookDetails[0].value = "Success"
        else:
            ebookDetails[0].value = "Failed"
            books_count += 1
            continue

        # Transform size and convert to epub, png, tiff file
        with yaspin(text=f'Converting page size to {ebookDetails[3].value}...', color="cyan") as sp:
            # Transform size 1 inch = 72 points
            width = float(ebookDetails[3].value.split(' x ')[0])
            height = float(ebookDetails[3].value.split(' x ')[1])
            width *= inch
            height *= inch
            
            coverpage_number = int(ebookDetails[4].value)
            firstpage_number = int(ebookDetails[5].value)
            
            reader = PdfReader(f'download/{ebookDetails[1].value}/tmp_{ebookDetails[1].value}.pdf')
            writer = PdfWriter()
            writer_for_coverpage = PdfWriter()
                        
            for i, page in enumerate(reader.pages):
                origin_height = float(page.cropbox.height)
                origin_width = float(page.cropbox.width)
                page.scale(width / origin_width, height / origin_height)
                if i + 1 == coverpage_number:                
                    writer_for_coverpage.add_page(page)
                if i + 1 < firstpage_number:
                    continue
                writer.add_page(page)
            writer.write(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_manuscript.pdf')
            writer_for_coverpage.write(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_coverpage.pdf')
            os.remove(f'download/{ebookDetails[1].value}/tmp_{ebookDetails[1].value}.pdf')
            sp.write(f"✅ Page size conversion complete!")
        
        # Make a coverpage png and tiff file
        with yaspin(text=f'Making coverpage images of {ebookDetails[1].value}...', color="cyan") as sp:
            images = convert_from_path(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_coverpage.pdf', dpi = 72, use_cropbox=True, transparent= True, use_pdftocairo=True)
            images[0].save(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_coverpage.png')
            
            img = Image.open(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_coverpage.png')
            img = img.resize((int(COVERPAGE_WIDTH), int(COVERPAGE_HEIGHT)))
            
            img.save(f'download/{ebookDetails[1].value}/{ebookDetails[1].value}_coverpage.tiff')
                        
            sp.write(f"✅ Finished making coverpage images of {ebookDetails[3].value}!")
        
        # Convert epub
        with yaspin(text=f'Converting {ebookDetails[1].value}_manuscript.pdf to {ebookDetails[1].value}_manuscript.epub...', color="cyan") as sp:
            os.system(f'ebook-convert "download/{ebookDetails[1].value}/{ebookDetails[1].value}_manuscript.pdf" "download/{ebookDetails[1].value}/{ebookDetails[1].value}_manuscript.epub"')
            sp.write(f"✅ Converted {ebookDetails[1].value}_manuscript.pdf to {ebookDetails[1].value}_manuscript.epub!")
        
        # Save xls file
        try:
            with yaspin(text="Saving Changes in your Excel File...", color = "green") as sp:
                # Saving changes to Excel File
                workbook.save(filename=Excelfilepath)

                #saves changes in the Excel sheet
                workbook.close()
                sp.write(f"✅ Updated the excel file!")
        except:
            Tk().withdraw()
            messagebox.showerror(title="Excel Edit Error",message="Unable to Edit the excel file\nFile is either:\n1) In Read Only mode\n2)Is corrupted\n3) Is already open")
    
        books_count += 1
    # Download pdf file and save

    # download_file(url, "1")
